import openpyxl

def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)


def write_excel_xlsx_append(path, sheet_name, value):
    index = len(value) 
    xfile = openpyxl.load_workbook(path)
    sheet = xfile.get_sheet_by_name(sheet_name)
    rows_old = sheet.max_row
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1+rows_old, column=j+1, value=str(value[i][j]))
    xfile.save(path)


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

